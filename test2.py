import pandas as pd
import numpy as np
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
import sending_email

sending_mail=False
email_subject="🟡Golden Cross in the last 7 trading days"
email_body=""

windows = {
    # "window1": [5, 20],
    "window2": [20, 50],
    "window3": [50, 200]
}

# --- CONFIG ---
excel_path = "combined_excel.xlsx"

def detect_golden_cross(excel_path, short_window, long_window, recent_window=7):
    if long_window >= 20 and long_window<50:
        lookback_days = long_window + 30
    elif long_window >= 50 and long_window<200:
        lookback_days = long_window + 50
    else:
        lookback_days = long_window + 100

    wb = load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    selected_sheets = sheet_names[:lookback_days]

    def read_sheet(sheet_name):
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=["Symbol", "Close"], engine="openpyxl")
            df.columns = df.columns.str.strip()
            df['Date'] = pd.to_datetime(sheet_name, format="%Y_%m_%d")
            df['Close'] = pd.to_numeric(df['Close'], errors='coerce')
            return df.dropna(subset=['Symbol', 'Close'])
        except Exception as e:
            print(f"⚠️ Skipping {sheet_name}: {e}")
            return pd.DataFrame()

    with ThreadPoolExecutor() as executor:
        dfs = list(executor.map(read_sheet, selected_sheets))

    combined_df = pd.concat(dfs, ignore_index=True)
    combined_df.sort_values(['Symbol', 'Date'], inplace=True)
    combined_df['Prev_Close'] = combined_df.groupby('Symbol')['Close'].shift(1)
    combined_df = combined_df[(combined_df['Close'] != combined_df['Prev_Close']) | (combined_df['Prev_Close'].isna())]
    combined_df.drop(columns=['Prev_Close'], inplace=True)

    start_date = combined_df['Date'].min().date()
    end_date = combined_df['Date'].max().date()

    results = []
    symbols = combined_df['Symbol'].unique()

    for symbol in symbols:
        df = combined_df[combined_df['Symbol'] == symbol].copy()
        df = df.sort_values('Date')
        df[f'SMA{short_window}'] = df['Close'].rolling(window=short_window, min_periods=1).mean()
        df[f'SMA{long_window}'] = df['Close'].rolling(window=long_window, min_periods=1).mean()
        df['GoldenCross'] = (
            df[f'SMA{short_window}'] > df[f'SMA{long_window}']
        ) & (
            df[f'SMA{short_window}'].shift(1) <= df[f'SMA{long_window}'].shift(1)
        )
        df['Symbol'] = symbol
        results.append(df)

    processed = pd.concat(results, ignore_index=True)
    unique_dates = processed['Date'].drop_duplicates().sort_values(ascending=False)
    recent_unique_dates = unique_dates[:recent_window]

    recent_crosses = processed[
        (processed['Date'].isin(recent_unique_dates)) & (processed['GoldenCross'])
    ]

    global email_body
    output=f"\n🟡 [{short_window}-{long_window}] Golden Cross detected in last {recent_window} trading days (up to {recent_unique_dates.max().date()}):"
    print(output)
    email_body+=output
    
    if recent_crosses.empty:
        print("No Golden Cross detected.")
    else:
        sorted_crosses = recent_crosses.sort_values('Date', ascending=False)
        for _, row in sorted_crosses.iterrows():
            output=f" - {row['Symbol']}: {row['Date'].date()}"
            print(output)
            email_body+=output



if __name__ == "__main__":
    for label, (short_window, long_window) in windows.items():
        detect_golden_cross(excel_path, short_window, long_window, recent_window=7)
    if len(email_body)==0:
        sending_mail=False
    if sending_mail:
        sending_email.send_email(email_subject,email_body)

