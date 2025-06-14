import pandas as pd
import numpy as np
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor

# --- CONFIG ---
excel_path = "combined_excel.xlsx"
short_window = 20
long_window = 50
if(long_window==20):
    lookback_days = long_window + 30
else:
    lookback_days = long_window + 50
recent_window = 7

# --- STEP 1: Read latest N sheets ---
wb = load_workbook(excel_path, read_only=True)
sheet_names = wb.sheetnames  # Ordered from newest to oldest
selected_sheets = sheet_names[:lookback_days]

def read_sheet(sheet_name):
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=["Symbol", "Close"], engine="openpyxl")
        df.columns = df.columns.str.strip()
        df['Date'] = pd.to_datetime(sheet_name, format="%Y_%m_%d")
        df['Close'] = pd.to_numeric(df['Close'], errors='coerce')
        return df.dropna(subset=['Symbol', 'Close'])
    except Exception as e:
        print(f"⚠️ Skipping {sheet_name}: {e}")
        return pd.DataFrame()

# --- STEP 2: Read all selected sheets in parallel ---
with ThreadPoolExecutor() as executor:
    dfs = list(executor.map(read_sheet, selected_sheets))

combined_df = pd.concat(dfs, ignore_index=True)
combined_df.sort_values(['Symbol', 'Date'], inplace=True)

# --- STEP 3: Remove duplicate closes (holidays, etc.) ---
combined_df['Prev_Close'] = combined_df.groupby('Symbol')['Close'].shift(1)
combined_df = combined_df[(combined_df['Close'] != combined_df['Prev_Close']) | (combined_df['Prev_Close'].isna())]
combined_df.drop(columns=['Prev_Close'], inplace=True)

start_date = combined_df['Date'].min().date()
end_date = combined_df['Date'].max().date()
# print(f"\n📅 Data range used for Golden Cross calculation: {start_date} to {end_date}")

# --- STEP 4: Detect golden cross manually without groupby().apply() ---
results = []

symbols = combined_df['Symbol'].unique()

for symbol in symbols:
    df = combined_df[combined_df['Symbol'] == symbol].copy()
    df = df.sort_values('Date')
    
    

    df[f'SMA{short_window}'] = df['Close'].rolling(window=short_window, min_periods=1).mean()
    df[f'SMA{long_window}'] = df['Close'].rolling(window=long_window, min_periods=1).mean()

    df['GoldenCross'] = (
        df[f'SMA{short_window}'] > df[f'SMA{long_window}']
    ) & (
        df[f'SMA{short_window}'].shift(1) <= df[f'SMA{long_window}'].shift(1)
    )

    df['Symbol'] = symbol
    results.append(df)

processed = pd.concat(results, ignore_index=True)

# --- STEP 5: Find golden crosses in recent N unique trading days ---
unique_dates = processed['Date'].drop_duplicates().sort_values(ascending=False)
recent_unique_dates = unique_dates[:recent_window]

recent_crosses = processed[
    (processed['Date'].isin(recent_unique_dates)) & (processed['GoldenCross'])
]

# --- STEP 6: Output ---
golden_symbols = recent_crosses['Symbol'].unique()

print(f"\n🟡 Golden Cross detected in last {recent_window} unique trading days (up to {recent_unique_dates.max().date()}):")

if recent_crosses.empty:
    print("No Golden Cross detected in the recent window.")
else:
    print("\n🟡 Golden Cross formed on these dates:")
    # Sort by Date descending before printing
    sorted_crosses = recent_crosses.sort_values('Date', ascending=False)

    for _, row in sorted_crosses.iterrows():
        print(f" - {row['Symbol']}: {row['Date'].date()}")
